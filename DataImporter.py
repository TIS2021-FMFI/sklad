from openpyxl import *
from Warehouse import Warehouse
from ShelvingUnit import *
from state1 import State
import time


class DataImporter:

    def __init__(self):
        self.__global_path = "./data/export.XLSX"
        self.__warehouse_type = "121"

    def set_data_file(self, path: str):
        self.__global_path = path

    def set_warehouse_type(self, warehouse_type: str):
        self.__warehouse_type = warehouse_type

    def get_shelving_units(self, warehouse: Warehouse):
        try:
            self.update_warehouse(warehouse)
        except PermissionError:
            try:
                time.sleep(5)
                self.update_warehouse(warehouse)
            except PermissionError:
                return 

    def update_warehouse(self, warehouse: Warehouse):
        wb = load_workbook(self.__global_path)
        ws = wb["Sheet1"]
        warehouse_types = ws['A']
        warehouse_places = ws['B']
        materials = ws['C']
        for i in range(1, len(warehouse_places)):
            if warehouse_types[i].value != self.__warehouse_type:
                continue
            names = warehouse_places[i].value.split('-')
            if len(names) < 3 or len(names) > 4:
                continue
            if len(names) == 3:
                shelvingunit_name, shelf_name, cell_name = names
            else:
                shelvingunit_name = names[0]
                shelf_name = names[1]
                cell_name = names[2]
            try:
                warehouse[shelvingunit_name][shelf_name][cell_name]
            except KeyError:
                continue
            material = materials[i].value
            if material != "<< prázdny >>":
                cell = warehouse[shelvingunit_name][shelf_name][cell_name]
                cell.set_state(State.OCCUPIED)
                if cell.is_blocked():
                    cell.change_block_state()                   
            else:
                warehouse[shelvingunit_name][shelf_name][cell_name].set_state(State.FREE)
